from __future__ import print_function
import os
import argparse
import base64
import json
import glob
import datetime
import logging
from logging.handlers import RotatingFileHandler
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# If modifying these SCOPES, delete the token.json files.
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

TOKEN_DIR = 'tokens'
OUTPUT_DIR = 'output'

def setup_directories():
    if not os.path.exists(TOKEN_DIR):
        os.makedirs(TOKEN_DIR)
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

def setup_logging():
    log_dir = 'logs'
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # File handler (rotates at 5MB, keeps 3 backups)
    fh = RotatingFileHandler(os.path.join(log_dir, 'audit.log'), maxBytes=5*1024*1024, backupCount=3)
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    
    # Console handler
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    logger.addHandler(ch)

def get_credentials(account_name):
    """Gets valid user credentials from storage or initiates OAuth2 flow."""
    creds = None
    token_path = os.path.join(TOKEN_DIR, f'{account_name}.json')
    
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
        
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists('credentials.json'):
                logging.error("credentials.json not found. Please download it from Google Cloud Console.")
                return None
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0) # port=0 picks an arbitrary available port
        # Save the credentials for the next run
        with open(token_path, 'w') as token:
            token.write(creds.to_json())
            
    return creds

def get_label_mapping(service):
    """Fetches user labels and returns a dictionary mapping labelId to label name."""
    try:
        results = service.users().labels().list(userId='me').execute()
        labels = results.get('labels', [])
        return {label['id']: label['name'] for label in labels}
    except HttpError as error:
        logging.error(f"An error occurred fetching labels: {error}")
        return {}

def extract_body(payload):
    """Recursively extracts the plain text body from the payload."""
    body = ""
    if 'parts' in payload:
        for part in payload['parts']:
            if part['mimeType'] == 'text/plain':
                data = part['body'].get('data', '')
                if data:
                    body += base64.urlsafe_b64decode(data).decode('utf-8')
            elif 'parts' in part:
                body += extract_body(part)
    elif payload['mimeType'] == 'text/plain':
        data = payload['body'].get('data', '')
        if data:
            body += base64.urlsafe_b64decode(data).decode('utf-8')
    return body

def classify_email(label_names):
    """Classifies email into 'Important' or 'Other' based on a consolidated rule set."""
    # 1. User explicitly marked it as important
    if 'IMPORTANT' in label_names or 'STARRED' in label_names:
        return 'Important'
        
    # 2. It's actively in the inbox
    if 'INBOX' in label_names:
        # If it's a promotion, social, or forum, it's noise
        if any(cat in label_names for cat in ['CATEGORY_PROMOTIONS', 'CATEGORY_SOCIAL', 'CATEGORY_FORUMS']):
            return 'Other'
        # Otherwise (Primary or Updates), we consider it Important
        return 'Important'
        
    # 3. If it's not in the inbox and not explicitly important, it's Other (Archived/Noise)
    return 'Other'

def load_priorities():
    """Loads priority mapping rules from priorities.json."""
    try:
        if os.path.exists('priorities.json'):
            with open('priorities.json', 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logging.warning(f"Could not load priorities.json: {e}")
    return []

def fetch_emails_for_account(account_name, after_date=None, before_date=None, max_results=10):
    """Fetches and processes emails for a specific account."""
    logging.info(f"--- Processing account: {account_name} ---")
    try:
        creds = get_credentials(account_name)
        if not creds:
            return
            
        service = build('gmail', 'v1', credentials=creds)
        
        # Get label mapping
        label_mapping = get_label_mapping(service)
        
        query_parts = []
        if after_date:
            query_parts.append(f"after:{after_date}")
        if before_date:
            query_parts.append(f"before:{before_date}")
        query = " ".join(query_parts)
            
        results = service.users().messages().list(
            userId='me',
            maxResults=max_results,
            q=query
        ).execute()
        
        messages = results.get('messages', [])
        if not messages:
            logging.info("No messages found.")
            return
            
        # Dictionary to hold messages grouped by date received and category
        messages_by_date_and_category = {}
            
        logging.info(f"Fetching details for {len(messages)} messages...")
        for msg_info in messages:
            msg_data = service.users().messages().get(userId='me', id=msg_info['id'], format='full').execute()
            payload = msg_data['payload']
            headers = payload.get('headers', [])
            
            # Extract headers
            subject = next((h['value'] for h in headers if h['name'].lower() == 'subject'), "")
            to = next((h['value'] for h in headers if h['name'].lower() == 'to'), "")
            cc = next((h['value'] for h in headers if h['name'].lower() == 'cc'), "")
            bcc = next((h['value'] for h in headers if h['name'].lower() == 'bcc'), "")
            sender = next((h['value'] for h in headers if h['name'].lower() == 'from'), "")
            
            # Extract body and get first 5 lines
            full_body = extract_body(payload)
            body_lines = full_body.strip().split('\n')
            summary = '\n'.join(body_lines[:5]).strip()
            if not summary:
                # Fallback to snippet if body extraction failed
                summary = msg_data.get('snippet', '')
                
            # Extract labels
            label_ids = msg_data.get('labelIds', [])
            label_names = [label_mapping.get(label_id, label_id) for label_id in label_ids]
            labels_str = ", ".join(label_names)
            
            # Extract timestamp
            internal_date = int(msg_data.get('internalDate', 0))
            if internal_date > 0:
                dt = datetime.datetime.fromtimestamp(internal_date / 1000.0)
                received_timestamp = dt.strftime('%Y-%m-%d %H:%M:%S')
                date_key = dt.strftime('%Y%m%d')
            else:
                received_timestamp = ""
                date_key = datetime.datetime.now().strftime('%Y%m%d')
                
            # Link to open mail
            mail_link = f"https://mail.google.com/mail/u/0/#all/{msg_info['id']}"
            
            row_data = {
                'Labels': labels_str,
                'From': sender,
                'Subject': subject,
                'Timestamp': received_timestamp,
                'To': to,
                'Cc': cc,
                'Bcc': bcc,
                'Summary': summary,
                'Id': msg_info['id'],
                'Link': mail_link
            }
            
            category = classify_email(label_names)
            group_key = (date_key, category)
            
            if group_key not in messages_by_date_and_category:
                messages_by_date_and_category[group_key] = []
            messages_by_date_and_category[group_key].append(row_data)
                
        # Write out to Excel files
        priorities = load_priorities()
        bold_font = Font(bold=True)
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
        for (date_key, category), records in messages_by_date_and_category.items():
            if not records:
                continue
                
            # Sort newest on top
            records.sort(key=lambda x: x['Timestamp'], reverse=True)
                
            filename = f"{account_name}_{category}_{date_key}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Emails"
            
            headers_list = ['Labels', 'From', 'Subject', 'Timestamp', 'To', 'Cc', 'Bcc', 'Summary', 'Id', 'Link', 'PriorityMatch']
            ws.append(headers_list)
            
            for row_idx, record in enumerate(records, start=2): # Row 1 is header
                is_priority = False
                for rule in priorities:
                    for k, v in rule.items():
                        if k.lower() == 'from' and v.lower() in record['From'].lower():
                            is_priority = True
                        if k.lower() == 'subject' and v.lower() in record['Subject'].lower():
                            is_priority = True
                
                record['PriorityMatch'] = 'Yes' if is_priority else 'No'
                
                row = [record.get(h, '') for h in headers_list]
                ws.append(row)
                
                if is_priority:
                    # Highlight entire row
                    for col_idx in range(1, len(headers_list) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = bold_font
                        cell.fill = yellow_fill
                        
            wb.save(filepath)
            logging.info(f"Saved {len(records)} records to {filepath}")
            
    except HttpError as error:
        logging.error(f"An error occurred: {error}")
    except Exception as e:
        logging.exception(f"An unexpected error occurred while processing account {account_name}: {e}")

def main():
    global TOKEN_DIR, OUTPUT_DIR
    
    parser = argparse.ArgumentParser(description="Fetch and categorize Gmail emails.")
    parser.add_argument("--add-account", help="Authenticate a new account and save its token.", type=str, metavar="EMAIL")
    parser.add_argument("--remove-account", help="Remove an authorized account token.", type=str, metavar="EMAIL")
    parser.add_argument("--run-all", help="Run extraction for all saved accounts.", action="store_true")
    parser.add_argument("--after", help="Fetch emails after this date (format YYYY/MM/DD)", required=False)
    parser.add_argument("--before", help="Fetch emails before this date (format YYYY/MM/DD)", required=False)
    parser.add_argument("--today", help="Fetch only today's emails. Overrides --after.", action="store_true")
    parser.add_argument("--max", help="Maximum results per account", type=int, default=10)
    parser.add_argument("--token-dir", help="Directory to load/save tokens.", default='tokens')
    parser.add_argument("--output-dir", help="Directory to save output files.", default='output')
    args = parser.parse_args()

    TOKEN_DIR = args.token_dir
    OUTPUT_DIR = args.output_dir

    setup_directories()
    setup_logging()

    if args.add_account:
        logging.info(f"Authenticating account: {args.add_account}")
        get_credentials(args.add_account)
        logging.info(f"Token saved successfully for {args.add_account}.")
        return

    if args.remove_account:
        token_path = os.path.join(TOKEN_DIR, f"{args.remove_account}.json")
        if os.path.exists(token_path):
            os.remove(token_path)
            logging.info(f"Successfully removed account: {args.remove_account}")
        else:
            logging.warning(f"Account not found: {args.remove_account}")
        return

    if args.run_all:
        token_files = glob.glob(os.path.join(TOKEN_DIR, '*.json'))
        if not token_files:
            logging.warning("No accounts found. Please use --add-account <email> to authorize an account first.")
            return
            
        for token_file in token_files:
            account_name = os.path.basename(token_file).replace('.json', '')
            
            fetch_after = args.after
            if args.today:
                fetch_after = datetime.datetime.now().strftime('%Y/%m/%d')
                
            fetch_emails_for_account(account_name, after_date=fetch_after, before_date=args.before, max_results=args.max)
    else:
        logging.warning("Please specify an action. Use --add-account <email>, --remove-account <email>, or --run-all. Use -h for help.")

if __name__ == '__main__':
    main()
